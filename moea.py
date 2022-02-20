import pandas as pd
import random
import json
import networkx as nx
import numpy as np
import itertools
import copy
import matplotlib.pyplot as plt
from collections import Counter
import math
from sklearn import metrics
import onmi
import omega_index
from Omega import Omega
import demon as d

from openpyxl import load_workbook


def write(wb1, com_node, fitness, col, off, onmi, omega):
    wb1.cell(row=1, column=1 + off * 4, value=fitness)
    wb1.cell(row=1, column=2 + off * 4, value=onmi)
    wb1.cell(row=1, column=3 + off * 4, value=omega)
    for com, nodes in com_node.items():
        # count=2
        wb1.cell(row=com + 2, column=1 + off * 4, value=str(nodes))


def write_all(individual_list, onmi_list, omega):
    wb = load_workbook("data.xlsx")
    wb1 = wb.active
    for i in range(len(individual_list)):
        write(wb1, individual_list[i].get_com_node(), individual_list[i].get_fitness(), i + 1, i, onmi_list[i],
              omega[i])
    wb.save("data.xlsx")


def cal_dissimilarity(G, node_num):
    d = dict(nx.algorithms.shortest_paths.weighted.all_pairs_dijkstra_path_length(G))
    theta = []
    for i in range(node_num):
        theta.append({j: 0 for j in range(node_num)})
    for i in range(node_num):
        for j in range(node_num):
            sum_ij = 0
            if i < j:
                for k in range(node_num):
                    if k != i and k != j:
                        sum_ij = sum_ij + (d[i][k] - d[j][k]) * (d[i][k] - d[j][k])
                theta[i][j] = math.sqrt(sum_ij) / (node_num - 2)
                theta[j][i] = theta[i][j]
    return theta


def clean_noise(must_link, cannot_link, theta):
    for ml in must_link:
        if theta[ml[0]][ml[1]] > 0.14:
            must_link.remove(ml)

    for nl in cannot_link:
        if theta[nl[0]][nl[1]] < 0.14:
            cannot_link.remove(nl)
    return must_link, cannot_link


def index_of(a, list):
    for i in range(0, len(list)):
        if list[i] == a:
            return i
    return -1


def sort_by_values(list1, values):
    sorted_list = []
    while (len(sorted_list) != len(list1)):
        if index_of(min(values), values) in list1:
            sorted_list.append(index_of(min(values), values))
        values[index_of(min(values), values)] = math.inf
    return sorted_list


def crowding_distance(values1, values2, front):
    distance = [0 for i in range(0, len(front))]
    sorted1 = sort_by_values(front, values1[:])
    sorted2 = sort_by_values(front, values2[:])
    distance[0] = 4444444444444444
    distance[len(front) - 1] = 4444444444444444
    for k in range(1, len(front) - 1):
        distance[k] = distance[k] + (values1[sorted1[k + 1]] - values2[sorted1[k - 1]]) / (max(values1) - min(values1))
    for k in range(1, len(front) - 1):
        distance[k] = distance[k] + (values1[sorted2[k + 1]] - values2[sorted2[k - 1]]) / (max(values2) - min(values2))
    return distance


def fast_non_dominated_sort(values1, values2):
    S = [[] for i in range(0, len(values1))]
    front = [[]]
    n = [0 for i in range(0, len(values1))]
    rank = [0 for i in range(0, len(values1))]

    for p in range(0, len(values1)):
        S[p] = []
        n[p] = 0
        for q in range(0, len(values1)):
            if (values1[p] > values1[q] and values2[p] > values2[q]) or (
                    values1[p] >= values1[q] and values2[p] > values2[q]) or (
                    values1[p] > values1[q] and values2[p] >= values2[q]):
                if q not in S[p]:
                    S[p].append(q)
            elif (values1[q] > values1[p] and values2[q] > values2[p]) or (
                    values1[q] >= values1[p] and values2[q] > values2[p]) or (
                    values1[q] > values1[p] and values2[q] >= values2[p]):
                n[p] = n[p] + 1
        if n[p] == 0:
            rank[p] = 0
            if p not in front[0]:
                front[0].append(p)

    i = 0
    while (front[i] != []):
        Q = []
        for p in front[i]:
            for q in S[p]:
                n[q] = n[q] - 1
                if (n[q] == 0):
                    rank[q] = i + 1
                    if q not in Q:
                        Q.append(q)
        i = i + 1
        front.append(Q)

    del front[len(front) - 1]
    return front


def cal_fitness2(must_link, cannot_link, node_com, centrality_dict):
    value_must_link = 0
    value_cannot_link = 0
    for e in must_link:
        # in the same community: reward
        if node_com[e[0]] == node_com[e[1]]:
            value_must_link = value_must_link + centrality_dict[e[0]] + centrality_dict[e[1]]
        else:
            pass  # add punishment
    for e in cannot_link:
        # in the same community but actually should not: punishment
        if node_com[e[0]] == node_com[e[1]]:
            value_cannot_link = value_cannot_link + centrality_dict[e[0]] + centrality_dict[e[1]]
        else:
            pass  # add reward
    return value_must_link - value_cannot_link


def after_decoding_2(genes, com_index_temp, neighbor_dict_temp):
    neighbor_dict = copy.deepcopy(neighbor_dict_temp)
    for com, edges in com_index_temp.items():
        if len(edges) < 4:
            # reassign
            for e in edges:
                if e in neighbor_dict[e]:
                    neighbor_dict[e].remove(e)
                new_value = np.random.choice(neighbor_dict[e])
                if new_value == genes[e]:
                    genes[e] = np.random.choice(neighbor_dict[e])
                else:
                    genes[e] = new_value
    return genes


def after_decoding(genes, com_index_temp, neighbor_dict, community_index, community_edge):
    com_index = copy.deepcopy(community_index)
    for com, edges in com_index_temp.items():
        if len(edges) < 4:
            # reassign
            for e in edges:
                selection_elements = list(set(com_index[community_edge[e]]).intersection(set(neighbor_dict[e])))
                selection_elements.remove(e)
                if len(selection_elements) == 0:
                    genes[e] = np.random.choice(neighbor_dict[e])
                else:
                    genes[e] = np.random.choice(selection_elements)
    return genes


# degree_dict: node: degree
def calc_EQ(m, degree_dict, community_node, A_org, node_community):
    EQ = 0
    A = copy.deepcopy(A_org)
    for nodes in community_node.values():
        for i in nodes:
            for j in nodes:
                # if A[i][j] != -1: and i != j
                # if i > j:
                #     continue
                # else:
                o_i = len(node_community[i])
                o_j = len(node_community[j])
                EQ = EQ + (A[i][j] - degree_dict[i] * degree_dict[j] / 2 / m) / o_i / o_j
                # A[i][j] = -1
                # A[j][i] = -1
    return EQ / 2 / m


def draw(node_list, edge_list, individual):
    G_final = nx.Graph()
    G_final.add_nodes_from(node_list)
    G_final.add_edges_from(edge_list)
    colors = [individual.get_community()[get_edge_index(edge, edge_list)] for edge in edge_list]
    pos = nx.spring_layout(G_final)
    nodes = nx.draw_networkx_nodes(G_final, pos, node_color='r', nodelist=node_list, node_size=30)
    edges = nx.draw_networkx_edges(G_final, pos=pos, edge_color=colors, width=3, edge_cmap=plt.cm.Blues,
                                   edgelist=edge_list)
    print("fitness", individual.get_fitness())
    plt.show()


def com_edge_to_node(node_num, com, edge_list):
    # com: [key: community index, value: edges]
    # print("best", com)
    community_node = {node: [] for node in range(node_num)}
    for com_index, edges in com.items():
        for edge_index in edges:
            e = edge_list[edge_index]
            community_node[e[0]].append(com_index)
            community_node[e[1]].append(com_index)
    for node, communities in community_node.items():
        community_node[node] = list(set(communities))
    return community_node


def com_node_to_node_com(node_num, com_node):
    node_com = {node: [] for node in range(node_num)}
    for com, nodes in com_node.items():
        for n in nodes:
            node_com[n].append(com)
    return node_com


def edge_com_to_com_edge(edge_com):
    com_edge = {}
    for edge, com in edge_com.items():
        if com in com_edge.keys():
            com_edge[com].append(edge)
        else:
            com_edge[com] = [edge]
    return com_edge


def node_com_to_com_node(node_com):
    com_node = {}
    for node, communities in node_com.items():
        for com in communities:
            if com in com_node.keys():
                com_node[com].append(node)
            else:
                com_node[com] = [node]
    return com_node


def get_lfr():
    lfr_1 = pd.read_table('./lfr/lfr_1.dat', header=None)
    edge_list = lfr_1.values.tolist()
    for i in range(len(edge_list)):
        edge_list[i][0] = edge_list[i][0] - 1
        edge_list[i][1] = edge_list[i][1] - 1
    names = ['index', 'community']
    ground_truth = pd.read_table('./lfr/community.dat', header=None, names=names)
    ground_truth_dict = {i: list(map(int, ground_truth['community'][i].split(" ")[:-1])) for i in
                         range(len(ground_truth))}
    node_num = len(ground_truth)
    return node_num, edge_list, ground_truth_dict


def network():
    path = './com-amazon.txt'
    data = pd.read_table(path, sep='\t', header=None)

    # data = pd.read_table(path, header=None)
    return data


def evaluate(ground_truth, communities):
    omega = omega_index.Omega(communities, ground_truth)
    print("omega", omega.omega_score)
    return omega.omega_score
    # metrics.normalized_mutual_info_score(ground_truth, com)


def draw_result(label_dict, G):
    color_list = ['skyblue', 'orange', 'orchid', 'lightcoral', 'aquamarine', 'hotpink', 'blueviolet', 'darkgrey',
                  'burlywood', 'c', 'yellow', 'skyblue', 'orange', 'orchid', 'lightcoral', 'aquamarine', 'hotpink',
                  'blueviolet', 'darkgrey',
                  'burlywood', 'c', 'yellow']
    values = []
    max_val = 4
    if max_val == 0:
        pos = nx.draw_planar(G, cmap=plt.get_cmap('jet'), node_color=values, node_size=260, with_labels=True)
        plt.show()
    else:
        for node in G.nodes:
            if len(label_dict[node]) == 1:
                values.append(color_list[label_dict[node][0]])
                # values.append(label_dict[node][0] / 7+0.1)
            elif len(label_dict[node]) == 0:
                values.append('pink')
            else:
                values.append('yellowgreen')
        # print(values)
        pos = nx.draw_kamada_kawai(G, cmap=plt.get_cmap("RdYlBu"), node_color=values, node_size=260, with_labels=True)
        plt.show()


def find_neighbor(edges, node_list):
    # fill in the neighbor_dict: nodes and their neighboring nodes
    neighbor_dict = {node: [] for node in node_list}
    '''
    for node in node_list:
        neighbors = []
        for e in range(len(edges)):
            ''''''
            if node == edges[e][0]:
                neighbors.append(edges[e][1])
            elif node == edges[e][1]:
                neighbors.append(edges[e][0])
        neighbor_dict[node] = neighbors
    '''
    return neighbor_dict


def find_neighbor_edge(edge_list, edge_index_dict):
    # find the edges connected with each node
    neighbor_edge_dict = {e: [] for e in range(len(edge_list))}
    index = 0
    for edge in edge_list:
        # the first node
        for e in edge_index_dict[edge[0]]:
            neighbor_edge_dict[index].append(e)
        for e in edge_index_dict[edge[1]]:
            if not e in neighbor_edge_dict[index]:
                neighbor_edge_dict[index].append(e)
        index = index + 1
        # print(neighbor_edge_dict)
        # neighbor_edge_dict[index].remove(edge[0])  # 删去本身
    return neighbor_edge_dict


def json_dumper(data, path):
    """
    Function to save a JSON to disk.
    :param data: Dictionary of cluster memberships.
    :param path: Path for dumping the JSON.
    """
    with open(path, 'w') as outfile:
        json.dump(data, outfile)


def label_propagation_community(G):
    communities_generator = list(nx.algorithms.community.label_propagation_communities(G))
    m = []
    for i in communities_generator:
        m.append(list(i))
    return m


# local expansion
def get_degree(edge_list):
    degree = {}
    for edge in edge_list:
        if edge[0] in degree.keys():
            degree[edge[0]] += 1
        else:
            degree[edge[0]] = 1
        if edge[1] in degree.keys():
            degree[edge[1]] += 1
        else:
            degree[edge[1]] = 1
    return degree


def get_node_of_max_degree(dict):
    # 1.choose from all max nodes randomly
    max_list = []
    max_val = max(dict.values())
    for key, value in dict.items():
        if value == max_val:
            max_list.append(key)
    return random.sample(max_list, 1)[0]


def get_node_of_max_centrality(centrality, nodes):
    max_list = []
    max_val = max(centrality.values())
    for key, value in centrality.items():
        if value == max_val:
            max_list.append(key)
    return random.sample(max_list, 1)[0]

def cal_eigenvector_centrality(nodes, edges):
    G = nx.Graph()
    G.add_nodes_from(nodes)
    G.add_edges_from(edges)
    centrality = nx.eigenvector_centrality(G)
    return centrality

def get_p1():
    p_range = [i for i in range(7)]
    return np.random.choice(p_range)


def get_parent_individual(agent_p1, agent_p2, index_p1, p):
    p_range = [i for i in range(7)]
    p_copy = p.copy()
    count = 0
    for fitness in p:
        if fitness < 0:
            p_range.remove(count)
        count = count + 1
    p_sum = sum(p_copy)
    p_copy = np.divide(p_copy, p_sum)
    if agent_p1 == agent_p2:
        choice = np.random.choice(p_range)
        if choice == index_p1:
            return get_parent_individual(agent_p1, agent_p2, index_p1, p)
        else:
            return choice
    else:
        choice = np.random.choice(p_range)
    return np.random.choice(p_range, p=p_copy.ravel())


def encode(m, community_dict, community_by_index, neighbor_edge_dict):
    """
    community of edges->locus value representation
    Iterate through each gene (edge), choose a neighbor of its community randomly and assign it as the gene value
    input: community dict (key:edge; value:community number)
    output: individual (locus value representation)
    g[0]:locus g[1]:gene value (neighbor)
    """
    genes = {e: -1 for e in range(m)}
    community_by_index_copy = copy.deepcopy(community_by_index)  # the list that stores origial elements
    community_elements = copy.deepcopy(community_by_index)  # the list for change
    print("encode")
    for edge in range(m):
        # get the index of an edge
        com_index = community_dict[edge]
        neighbor_elements = copy.deepcopy(neighbor_edge_dict[edge])

        selection_elements = list(set(community_elements[com_index]).intersection(set(neighbor_elements)))
        if edge in selection_elements:
            selection_elements.remove(edge)
        if len(selection_elements) == 0:  # the community index: community_dict[edge]
            selection_elements_0 = list(set(community_by_index_copy[com_index]).intersection(set(neighbor_elements)))
            if edge in selection_elements_0:
                selection_elements_0.remove(edge)
            if len(selection_elements_0) == 0:
                neighbor_elements.remove(edge)
                genes[edge] = random.sample(neighbor_elements, 1)[0]
            else:
                temp = random.sample(selection_elements_0, 1)[0]
                genes[edge] = temp
                if genes[temp] == edge:
                    genes[edge] = random.sample(selection_elements_0, 1)[0]
        else:
            # no more nodes for selection
            temp_value = random.sample(selection_elements, 1)[0]
            genes[edge] = temp_value
            if genes[temp_value] == edge:
                genes[edge] = random.sample(selection_elements, 1)[0]
                community_elements[com_index].remove(genes[edge])

            else:
                community_elements[com_index].remove(temp_value)
    return genes


def decode(genes, m):
    """
    locus value representation -> a dict that stores the community number of each edge
    若该边不在任意一个community，则新建一个community，否则加入现有的community
    input: individual (locus value representation)
    output: community dict (key:edge; value:community number)
    g[0]:locus g[1]:gene value (neighbor)
    """
    count = 0
    community_edge = {e: -1 for e in range(m)}  # key: edge number; value: the community of the edge
    community_by_index = {}  # key:community_num; value: the list that stores edges in this community
    for g0, g1 in genes.items():
        if community_edge[g0] != -1:
            # +- locus already belong to some community
            if community_edge[g1] == -1:
                # value does not have any assignment
                # assign the community of locus to value
                com = community_edge[g0]
                community_edge[g1] = com
                community_by_index[com].append(g1)
            # ++
            else:
                if community_edge[g0] == community_edge[g1]:
                    continue
                elif community_edge[g0] < community_edge[g1]:
                    c0 = community_edge[g0]
                    c1 = community_edge[g1]
                else:
                    c0 = community_edge[g1]
                    c1 = community_edge[g0]
                for component in community_by_index[c1]:
                    community_by_index[c0].append(component)
                    community_edge[component] = c0

                del community_by_index[c1]

        else:
            # -+
            if community_edge[g1] != -1:  # value is in an existing community
                com = community_edge[g1]
                community_edge[g0] = com
                community_by_index[com].append(g0)
            else:
                # --
                community_edge[g0] = count
                community_edge[g1] = count
                community_by_index[count] = [g0, g1]
                count = count + 1
    community_edge_temp = {e: -1 for e in range(m)}
    community_by_index_temp = {c: [] for c in range(len(community_by_index.keys()))}
    counter = 0
    for index, edges in community_by_index.items():
        community_by_index_temp[counter] = community_by_index[index]
        for edge in community_by_index_temp[counter]:
            community_edge_temp[edge] = counter
        counter = counter + 1

    return community_edge_temp, community_by_index_temp


def get_edge_index(edge, edge_list):
    for index_e in range(len(edge_list)):
        if edge[0] == edge_list[index_e][0] and edge[1] == edge_list[index_e][1]:
            return index_e


def get_edge(edge_num, node_num, edge_list):
    edge_dict = {node: [] for node in range(node_num)}  # edges connected to each node
    # initialization: get_edge()
    edge_index_dict = {node: [] for node in range(node_num)}  # indexes of edges connected to each node
    for k in range(edge_num):
        t = edge_list[k]
        edge_dict[t[0]].append(t)
        edge_dict[t[1]].append(t)
        edge_index_dict[t[0]].append(k)
        edge_index_dict[t[1]].append(k)
    return edge_dict, edge_index_dict

class LocalExpansion:
    def __init__(self, nodes, mode, neighbor_dict, select_dict):
        self.nodes = nodes.copy()
        self.community = {node: -1 for node in self.nodes}
        self.community_count = 0
        self.flag = True
        self.mode = mode
        self.neighbor_dict = copy.deepcopy(neighbor_dict)
        self.select_dict = copy.deepcopy(select_dict)
        self.count = 0

    def expand(self):
        # 1.choose the node to expand
        seed = get_node_of_max_degree(self.select_dict)
        del self.select_dict[seed]
        self.community[seed] = self.community_count
        self.count = self.count + 1
        for node in self.neighbor_dict[seed]:
            # this node has not been set
            if self.community[node] == -1:
                self.community[node] = self.community_count
                self.count = self.count + 1
                del self.select_dict[node]
        self.community_count += 1

    def local_expansion(self):
        for key, val in self.neighbor_dict.items():
            self.neighbor_dict[key].remove(key)

        while self.count < len(self.nodes):
            self.expand()

    def get_num_community(self):
        return self.community_count

    def get_community(self):
        return self.community


# coonvert community of nodes to community of edges
def node_to_edge(node_num, edges, community):
    # community: {node: community index of the node}
    community_edge = {e: 0 for e in range(len(edges))}
    community_index = {n: [] for n in range(node_num)}
    edge_index = 0
    for edge in edges:
        # check the two ends of one edge
        c1 = community[edge[0]]
        c2 = community[edge[1]]
        # same community
        if c1 == c2:
            community_edge[edge_index] = c1
            community_index[c1].append(edge_index)
        else:
            # different communities: choose one community randomly from the two communities\
            random_node = random.sample([c1, c2], 1)[0]
            community_edge[edge_index] = random_node
            community_index[random_node].append(edge_index)
        edge_index = edge_index + 1
    return community_edge, community_index


class LabelPropagator:
    def __init__(self, nodes, neighbor_dict):
        self.nodes = nodes.copy()
        self.labels = {node: node for node in self.nodes}
        self.label_count = len(set(self.labels.values()))
        self.flag = False
        self.neighbor_dict = neighbor_dict

    def choose_label(self, source):
        """
        Calculate the frequency of labels and choose one
        :param neighbors: Neighboring nodes.
        :param source: Source node.
        """
        # number of possible labels in total: number of nodes
        scores = {node: 0 for node in range(len(self.nodes))}  # key: label, value: count
        for neighbor in self.neighbor_dict[source]:
            neighbor_label = self.labels[neighbor]
            scores[neighbor_label] = scores[neighbor_label] + 1
        top = [key for key, val in scores.items() if val == max(scores.values())]
        return random.sample(top, 1)[0]

    def propagate(self):
        """
        Doing a propagation round.
        """
        random.seed()
        random.shuffle(self.nodes)
        for node in self.nodes:
            pick = self.choose_label(node)
            current_label = self.labels[node]
            self.labels[node] = pick
            if pick != current_label:
                self.flag = False

    def label_propagation(self):
        """
        Doing propagations until convergence or reaching time budget.
        """
        index = 0
        # index < self.rounds
        while not self.flag:
            index = index + 1
            self.flag = True
            print("\nLabel propagation round: " + str(index) + ".\n")
            self.propagate()
        print("")
        return self.labels

    def get_num_community(self):
        return self.community_count


def init(G, m, population, degree_dict, centrality_dict, node_list, edge_list, neighbor_dict,
         neighbor_edge_dict, A, size, must_link, cannot_link):
    individual_list = []
    for index in range(size):
        print("individual", index)
        temp_fitness = 0
        temp_fitness_index = 0
        individual = Individual()
        type_of_init = np.random.uniform(0, 1)
        # label propagation
        if type_of_init < 0.33:  # 0.33
            # get a community of nodes
            print("label propagator")
            community_dict = label_propagation_community(G)
            label_dict = {0: node for node in node_list}
            count = 0
            for community in community_dict:
                for node in community:
                    label_dict[node] = count
                count = count + 1

            community_edge, community_index = node_to_edge(len(node_list), edge_list, label_dict)

            for key in list(community_index.keys()):
                if not community_index.get(key):
                    del community_index[key]

            genes_temp = encode(m, community_edge, community_index, neighbor_edge_dict)
            community_edge_temp, community_by_index_temp = decode(genes_temp, m)
            genes = after_decoding(genes_temp, community_by_index_temp, neighbor_edge_dict, community_index,
                                   community_edge)
            community_edge_temp_2, community_by_index_temp_2 = decode(genes, m)
            individual.set_genes(genes)
            individual.set_community_edge(community_edge_temp_2)
            individual.set_community_by_index(community_by_index_temp_2)
            individual.set_type(0)

            node_com_1 = com_edge_to_node(len(node_list), community_by_index_temp_2, edge_list)
            com_node_1 = node_com_to_com_node(node_com_1)
            fitness = calc_EQ(m, degree_dict, com_node_1, A, node_com_1)
            fitness2 = cal_fitness2(must_link, cannot_link, node_com_1, centrality_dict)
            individual.set_com_node(com_node_1)
            individual.set_fitness(fitness)
            individual.set_fitness2(fitness2)

        # LocalExpansion
        elif type_of_init < 0.67:
            # get a community of nodes
            print("local expansion")
            if np.random.uniform(0, 1) < 0.7:
                localExapansion = LocalExpansion(node_list, 0, neighbor_dict, degree_dict)
                print("degree_dict")
            else:
                random_dict = {node: np.random.randint(0, 30) for node in range(len(node_list))}
                localExapansion = LocalExpansion(node_list, 0, neighbor_dict, random_dict)
                print("random")
            localExapansion.local_expansion()
            label_dict = localExapansion.get_community()
            print("community_index", label_dict)

            # convert community of nodes to community of edges (not formatted)
            community_edge, community_index = node_to_edge(len(node_list), edge_list, label_dict)
            # convert the community of edges to
            for key in list(community_index.keys()):
                if not community_index.get(key):
                    del community_index[key]

            genes_temp = encode(m, community_edge, community_index, neighbor_edge_dict)
            community_edge_temp, community_by_index_temp = decode(genes_temp, m)
            genes = after_decoding(genes_temp, community_by_index_temp, neighbor_edge_dict, community_index,
                                   community_edge)
            community_edge_temp_2, community_by_index_temp_2 = decode(genes, m)
            individual.set_genes(genes)
            individual.set_community_edge(community_edge_temp_2)
            individual.set_community_by_index(community_by_index_temp_2)
            node_com_1 = com_edge_to_node(len(node_list), community_by_index_temp_2, edge_list)
            com_node_1 = node_com_to_com_node(node_com_1)
            fitness = calc_EQ(m, degree_dict, com_node_1, A, node_com_1)
            fitness2 = cal_fitness2(must_link, cannot_link, node_com_1, centrality_dict)
            individual.set_type(1)
            individual.set_fitness(fitness)
            individual.set_fitness2(fitness2)
            individual.set_com_node(com_node_1)

        # eigenvector centrality
        else:
            print("eigenvector centrality")
            # get a community of nodes
            eigenvectorCentrality = LocalExpansion(node_list, 1, neighbor_dict, centrality_dict)
            print("created")
            eigenvectorCentrality.local_expansion()
            label_dict = eigenvectorCentrality.get_community()

            # convert community of nodes to community of edges (not formatted)
            community_edge, community_index = node_to_edge(len(node_list), edge_list, label_dict)
            for key in list(community_index.keys()):
                if not community_index.get(key):
                    del community_index[key]
            # convert the community of edges to
            genes_temp = encode(m, community_edge, community_index, neighbor_edge_dict)
            community_edge_temp, community_by_index_temp = decode(genes_temp, m)
            genes = after_decoding(genes_temp, community_by_index_temp, neighbor_edge_dict, community_index,
                                   community_edge)
            community_edge_temp_2, community_by_index_temp_2 = decode(genes, m)
            individual.set_genes(genes)
            individual.set_community_edge(community_edge_temp_2)
            individual.set_community_by_index(community_by_index_temp_2)
            node_com_1 = com_edge_to_node(len(node_list), community_by_index_temp_2, edge_list)
            com_node_1 = node_com_to_com_node(node_com_1)
            fitness = calc_EQ(m, degree_dict, com_node_1, A, node_com_1)
            fitness2 = cal_fitness2(must_link, cannot_link, node_com_1, centrality_dict)
            individual.set_type(2)
            individual.set_fitness(fitness)
            individual.set_fitness2(fitness2)
            individual.set_com_node(com_node_1)
        individual_list.append(individual)
    return individual_list


def update_l3(population):
    for i_agent in range(13):
        fitness_list = population.get_fitness_list()[i_agent * 7:i_agent * 7 + 7].copy()
        max_index = fitness_list.index(max(fitness_list))
        if max_index != 0:
            population.swap_individual(i_agent, 0, max_index)

def update_l2(population):
    for i in range(1, 4):
        agent_list = population.get_agent_list().copy()
        agent_fitness = agent_list[i][0].get_fitness()  # 1 2 3
        # get the 1, 2 and 3 th pocket value
        # agent:456 789 101112
        temp_list_456 = [agent_list[j][0].get_fitness() for j in range(1 + i * 3, 4 + i * 3)]
        max_456 = max(temp_list_456)
        if max_456 > agent_fitness:
            offset = temp_list_456.index(max_456)
            temp_l2_index = offset + i * 3 + 1
            population.swap(i, temp_l2_index)
    update_l3(population)


def update(population):
    # find the pocket individual
    # For each agent: pocket=individual of the greatest fitness in [individual]
    # For leader: leader pocket= individual of the greatest fitness

    # proceed layer by layer

    # each agent reserve the first element as pocket element

    # 5 6 7 / 2 // 8 9 10 / 3 // 11 12 13 /4  /// 0

    agent_list = population.get_agent_list().copy()

    temp_l1_index = 2
    temp_l1_fitness = agent_list[temp_l1_index - 1][0].get_fitness()

    # from each agent
    update_l3(population)
    update_l2(population)

    temp_list_123 = [population.get_agent_list()[j][0].get_fitness() for j in range(1, 4)]
    max_123 = max(temp_list_123)
    if max_123 > population.get_agent_list()[0][0].get_fitness():
        population.swap(temp_list_123.index(max_123) + 1, 0)
        update_l2(population)


def local_search(neighbor_edge, individual, m, edge_list, node_num, degree_dict, A):
    community = individual.get_community()
    community_by_index = individual.get_community_by_index()
    neighbor_edge_dict = copy.deepcopy(neighbor_edge)
    count = 0
    for g in individual.get_genes().values():
        # 是否将g[1]改成edge
        c1 = community[g]

        for edge in neighbor_edge_dict[g]:
            neighbor_edge_dict[g].remove(edge)
            c2 = community[edge]
            if c1 != c2:

                g_copy = copy.deepcopy(individual.genes)
                g_copy[count] = edge
                community_edge_new, community_by_index_new = decode(g_copy, m)
                node_com_new = com_edge_to_node(node_num, community_by_index_new, edge_list)

                com_node_new = node_com_to_com_node(node_com_new)

                fitness_new = calc_EQ(m, degree_dict, com_node_new, A, node_com_new)
                community_edge_temp, community_by_index_temp = decode(individual.get_genes(), m)
                node_com_org = com_edge_to_node(node_num, community_by_index_temp, edge_list)
                com_node_org = node_com_to_com_node(node_com_org)
                fitness = calc_EQ(m, degree_dict, com_node_org, A, node_com_org)
                if fitness_new - fitness > 0:  # the new setup is better than the original one
                    individual.set_gene_index(edge, count)
                    individual.set_fitness(fitness_new)
                    break

        count = count + 1
    return individual


def diversity(fitness_list):
    return np.std(fitness_list)


def crossover(m, cp, parent1, parent2):
    # parents are individuals
    offspring1 = {}
    offspring2 = {}
    for edge in range(m):
        if np.random.uniform(0, 1) > cp:
            offspring1[edge] = parent1.get_genes()[edge]
            offspring2[edge] = parent2.get_genes()[edge]
        else:
            offspring2[edge] = parent1.get_genes()[edge]
            offspring1[edge] = parent2.get_genes()[edge]
    return offspring1, offspring2


def mutate(edge_list, gene, mp, neighbor_dict, marginal_p, degree_dict, node_num, m, A):
    # change the value of one gene if random number<mp
    # domain of selection: neighbors of g[0]
    count = 0
    genes = copy.deepcopy(gene)
    neighbor_edge_dict = copy.deepcopy(neighbor_dict)
    # print(neighbor_edge_dict)
    for e in range(len(edge_list)):
        neighbor_edge_dict[e].remove(e)
        if np.random.uniform(0, 1) < mp:
            # print("e",e)
            # no neighbor
            if len(neighbor_edge_dict[e]) == 0:  # 这里neighbor_edge是不包含本身的情况
                continue
            else:
                # have neighbor: choose one from the neighbors
                genes[count] = random.sample(neighbor_edge_dict[e], 1)[0]
        count = count + 1
    marginal_genes = set(list(range(len(edge_list)))).difference(set(genes.values()))

    community_edge_old, community_by_index_old = decode(genes, m)
    node_com_old = com_edge_to_node(node_num, community_by_index_old, edge_list)
    com_node_old = node_com_to_com_node(node_com_old)
    fitness_old = calc_EQ(m, degree_dict, com_node_old, A, node_com_old)

    genes_temp = copy.deepcopy(genes)
    for g in marginal_genes:
        value_temp = genes_temp[g]
        if np.random.uniform(0, 1) < marginal_p:
            if len(neighbor_edge_dict[g]) == 0:  # 这里neighbor_edge是不包含本身的情况
                continue
            else:
                # have neighbor: choose one from the neighbors
                genes_temp[g] = random.sample(neighbor_edge_dict[g], 1)[0]
                community_edge_temp, community_by_index_temp = decode(genes_temp, m)
                node_com_org = com_edge_to_node(node_num, community_by_index_temp, edge_list)
                com_node_org = node_com_to_com_node(node_com_org)
                fitness_new = calc_EQ(m, degree_dict, com_node_org, A, node_com_org)
                if fitness_new < fitness_old:
                    genes_temp[g] = value_temp
    return genes_temp


class Population:
    def __init__(self, size):
        self.individual_list = []
        self.diversity = 0
        self.size = size
        self.fitness_list = [0 for x in range(size)]
        self.fitness_list2 = [0 for x in range(size)]
        # self.neighbor_dict

    def update_diversity(self):
        self.diversity = diversity(self.fitness_list)

    def get_diversity(self):
        return self.diversity

    def change_fitness(self, fitness, index):
        self.fitness_list[index] = fitness

    def set_fitness_list(self, fitness_list):
        self.fitness_list = fitness_list

    def update_fitness_list(self):
        self.fitness_list = []
        self.fitness_list2 = []
        for i in range(self.size):
            self.fitness_list.append(self.individual_list[i].get_fitness())
            self.fitness_list2.append(self.individual_list[i].get_fitness2())

    def get_fitness_list(self):
        return self.fitness_list

    def get_fitness_list2(self):
        return self.fitness_list2

    def get_individual_list(self):
        return self.individual_list

    def set_individual_list(self, individual_list):
        self.individual_list = individual_list

    def swap(self, a, b):
        self.individual_list[a], self.individual_list[b] = self.individual_list[b], self.individual_list[a]
        self.fitness_list[a], self.fitness_list[b] = self.fitness_list[b], self.fitness_list[a]
        self.fitness_list2[a], self.fitness_list2[b] = self.fitness_list2[b], self.fitness_list2[a]


def partition_density(mi, n):
    if n <= 2:
        return 0
    return (mi - n + 1) / (n - 2) / (n - 1)


def partition_density_whole(m, community_dict, edge_list):
    sum_temp = 0
    for com in community_dict.keys():
        m_i = len(community_dict[com])  # number of internal edges
        n = get_node_num_in_community(edge_list, community_dict[com])
        sum_temp = sum_temp + m_i * partition_density(m_i, n)
    return sum_temp * 2 / m


def delta_fitness(m, genes, count, edge, community_by_index):
    # set the gene value of edge to g
    g_copy = genes.copy()
    g_copy[count] = edge
    community_edge_new, community_by_index_new = decode(g_copy)
    fitness_new = partition_density_whole(m, community_by_index_new)
    if fitness_new - partition_density_whole(m, community_by_index) < 0:  # the new setup is worse than the original one
        return 0, fitness_new, community_edge_new, community_by_index_new
    else:
        return 1, fitness_new, community_edge_new, community_by_index_new


def discard_node(n, edge_list):
    e_temp = []
    for e in edge_list:
        if e[0] == n or e[1] == n:
            continue
        else:
            e_temp.append(e)
    return e_temp, len(e_temp)


def get_edge_in_community(edge_list, nodes):
    edges = []
    for i in nodes:
        for j in nodes:
            if i < j:
                if (i, j) in edge_list:
                    edges.append((i, j))
    return len(edges), edges


def get_node_num_in_community(edge_list, community):
    # edge index of this community
    community_flat = []
    # 1.get the group of edges
    for e in community:
        community_flat.append(edge_list[e])
    # 2.flatten
    community_flat = set(list(itertools.chain.from_iterable(community_flat)))
    # 3.set
    elements = set(community_flat)
    return len(elements), elements


class Individual:
    def __init__(self, genes=[], community_edge={}, community_by_index=[], com_node=[]):
        self.fitness = -1
        self.fitness2 = -1
        self.type = 0
        self.index = 0
        self.genes = genes
        self.community_edge = community_edge  # key:edge; value: index
        self.community_by_index = community_by_index
        self.com_node = com_node

    def set_com_node(self, com_node):
        self.com_node = com_node

    def get_com_node(self):
        return self.com_node

    def update_individual(self, flg, genes, community_edge, community_by_index, edge_list, m, node_list, A,
                          degree_dict):
        self.community_edge = community_edge
        self.community_by_index = community_by_index
        if flg:
            self.genes = genes
            node_com_1 = com_edge_to_node(len(node_list), community_by_index, edge_list)
            com_node_1 = node_com_to_com_node(node_com_1)
            fitness = calc_EQ(m, degree_dict, com_node_1, A, node_com_1)
            self.fitness = fitness

    def set_type(self, tp):
        self.type = tp

    def set_gene_index(self, val, index):
        self.genes[index] = val

    def set_genes(self, genes):
        self.genes = genes

    def set_community_edge(self, community_edge):
        self.community_edge = community_edge

    def set_community_by_index(self, community_by_index):
        self.community_by_index = community_by_index

    def set_fitness(self, fitness):
        self.fitness = fitness

    def set_fitness2(self, fitness2):
        self.fitness2 = fitness2

    def set_index(self, index):
        self.index = index

    def get_index(self, index):
        return self.index

    def get_fitness(self):
        return self.fitness

    def get_fitness2(self):
        return self.fitness2

    def get_genes(self):
        return self.genes

    def get_community(self):
        return self.community_edge

    def get_community_by_index(self):
        return self.community_by_index


def ad(m, n):
    return 2 * m / n


def post_processing(node_com, com_node, m, degree_dict, A, G, edge_list):
    node_com_temp = copy.deepcopy(node_com)
    for node in node_com.keys():
        communities = copy.deepcopy(node_com[node])
        # nodes with multiple memberships
        if len(communities) > 1:
            # add this node to this community
            fitness_loss = 10
            com_index = -1
            flag = True
            # sort the communities according to the frequency
            neighboring_labels = []
            count_n = 0
            for neighbor in G.neighbors(node):
                neighboring_labels.extend(node_com_temp[neighbor])
            result = Counter(neighboring_labels)
            results = []
            for res in list(result.keys()):
                if res in communities:
                    results.append(res)
            rank = 0
            for community in communities:
                m_i, edges = get_edge_in_community(edge_list, com_node[community])
                keep = ad(m_i, len(com_node[community]))
                # discard
                nodes_discard = copy.deepcopy(com_node[community])
                nodes_discard.remove(node)
                edge_discard, m_discard = discard_node(node, edges)
                G_temp = nx.Graph()
                G_temp.add_edges_from(edge_discard)
                G_temp.add_nodes_from(nodes_discard)
                discard = ad(m_discard, len(com_node[community]) - 1)
                if keep < 1 * discard: #math.pow(1.1, rank+1)
                    # update when fitness loss is smaller than the stored fitness loss
                    if discard - keep < fitness_loss:
                        fitness_loss = discard - keep
                        com_index = community
                    node_com_temp[node].remove(community)
                    com_node[community].remove(node)
                else:
                    # improvement, do nothing
                    flag = False
                rank = rank + 1
            # all negative
            if com_index != -1 and flag:  # no assignment improves the fitness, choose the community with least fitness loss
                node_com_temp[node].append(com_index)
                com_node[com_index].append(node)

    com_node_temp = copy.deepcopy(com_node)

    return node_com_temp, com_node_temp


def post_processing_2(node_com, com_node, m, degree_dict, A, G):
    node_com_temp = copy.deepcopy(node_com)
    for node in node_com.keys():
        communities = node_com[node]
        # nodes with multiple memberships
        if len(communities) > 1:
            # add this node to this community
            fitness_loss = 10
            com_index = -1
            flag = True
            for community in communities:
                # if all neighbors do not have this label,skip
                neighboring_label = []
                count_n = 0
                for neighbor in G.neighbors(node):
                    neighboring_label.extend(node_com_temp[neighbor])
                    if community in node_com_temp[neighbor]:
                        count_n += 1
                # nl=set(neighboring_label)
                # not this community
                if count_n / len(neighboring_label) < 0.3:
                    node_com_temp[node].remove(community)
                    com_node[community].remove(node)
                    com_index = community
                    continue

                # partition density without this node
                pd_keep = calc_EQ(m, degree_dict, com_node, A, node_com)
                # m, degree_dict, community_node, A_org, node_community
                node_com_copy = copy.deepcopy(node_com_temp)
                com_node_copy = copy.deepcopy(com_node)
                node_com_copy[node].remove(community)
                com_node_copy[community].remove(node)

                pd_discard = calc_EQ(m, degree_dict, com_node_copy, A, node_com_copy)

                if pd_keep < 1 * pd_discard:
                    # update when fitness loss is smaller than the stored fitness loss
                    if pd_discard - pd_keep < fitness_loss:
                        fitness_loss = pd_discard - pd_keep
                        com_index = community
                    node_com_temp[node].remove(community)
                    com_node[community].remove(node)
                else:
                    # improvement, do nothing
                    flag = False
            # all negative
            if com_index != -1 and flag:  # no assignment improves the fitness, choose the community with least fitness loss
                node_com_temp[node].append(com_index)
                com_node[com_index].append(node)

    com_node_temp = copy.deepcopy(com_node)
    return node_com_temp, com_node_temp


def gen_prior_noise(ni_p, link, node_num):
    noisy_link = []
    for i in range(node_num):
        for j in range(node_num):
            if [i, j] not in link:
                noisy_link.append([i, j])
        noisy_link.remove([i, i])
    noisy_index = random.sample(list(range(len(noisy_link))), int(ni_p * len(link)))
    return [noisy_link[i] for i in noisy_index]


def gen_noisy_input(ni_p, edge_list, node_num):
    noisy_link = []
    for i in range(node_num):
        for j in range(node_num):
            if (i, j) not in edge_list:
                noisy_link.append((i, j))
        noisy_link.remove((i, i))
    noisy_index = random.sample(list(range(len(noisy_link))), int(ni_p * len(edge_list)))
    return [noisy_link[i] for i in noisy_index]


def gen_noisy_prior_info(mu_p, cn_p, ni_p, edge_list, node_num, com_node, theta):
    must_link, cannot_link = gen_prior_info(mu_p, cn_p, edge_list, com_node)
    noise_must_link = gen_prior_noise(ni_p, must_link, node_num)
    print("noisy must")
    for link in noise_must_link:
        print(theta[link[0]][link[1]])

    must_link = must_link + noise_must_link
    noise_cannot_link = gen_prior_noise(ni_p, cannot_link, node_num)

    print("noisy cannot")
    for link in noise_cannot_link:
        print(theta[link[0]][link[1]])

    cannot_link = cannot_link + noise_cannot_link
    return must_link, cannot_link


def gen_prior_info(mu_p, cn_p, edge_list, com_node):
    edge_copy = copy.deepcopy(edge_list)
    count_must_link = 0
    count_cannot_link = 0
    must_link = []
    cannot_link = []
    while count_must_link < int(mu_p * len(edge_list)):
        i = random.choice(list(com_node.keys()))
        # choose two nodes from com_node[i]
        n1 = random.choice(com_node[i])
        n2 = random.choice(com_node[i])
        while n1 == n2:
            n2 = np.random.choice(com_node[i])
        must_link.append([n1, n2])
        count_must_link = count_must_link + 1
    while count_cannot_link < int(cn_p * len(edge_list)):
        i = random.choice(list(com_node.keys()))
        j = random.choice(list(com_node.keys()))
        while i == j:
            j = random.choice(list(com_node.keys()))
        # choose two nodes from com_node[i]
        n1 = random.choice(com_node[i])
        n2 = random.choice(com_node[j])
        cannot_link.append([n1, n2])
        count_cannot_link = count_cannot_link + 1
    return must_link, cannot_link


def get_ground_truth(ground_truth):
    ground_truth_index = []
    names = list(set(ground_truth))
    for n in ground_truth:
        ground_truth_index.append(names.index(n))
    return ground_truth_index


def main():
    ni_p = 0  # noise in input
    np_p = 0  # noise in prior knowledge
    mu_p = 0
    cn_p = 0
    # get graph
    # data = network()

    # # ---------karate
    G = nx.karate_club_graph()
    node_num = len(G.nodes)
    data = nx.to_pandas_edgelist(G)
    ary_data = np.array(data)  # np.ndarray()
    edge_list = ary_data.tolist()
    karate_ground_truth = [G.nodes[v]['club'] for v in G.nodes()]
    ground_truth_index = get_ground_truth(karate_ground_truth)
    ground_truth_dict = {i: [ground_truth_index[i]] for i in range(node_num)}
    ground_truth_com = node_com_to_com_node(ground_truth_dict)
    ground_truth_list = [set(v) for v in ground_truth_com.values()]
    edge_list_copy = copy.deepcopy(edge_list)
    noise_edge = gen_noisy_input(ni_p, edge_list, node_num)
    edge_list_copy.extend(noise_edge)
    print(type(edge_list_copy))
    G.add_edges_from(edge_list_copy)

    # for e in edge_list:
    #     with open("karate_cf.txt", "a") as f:
    #         f.write(str(e[0])+" "+str(e[1])+"\n")
    # theta = cal_dissimilarity(G, node_num)
    # print(ground_truth_com)
    # # ---------end karate

    # # -------------------football------------
    # G_org = nx.read_gml('polbooks.gml')
    # node_G = G_org.nodes
    # node_list_org = []
    # for node in node_G:
    #     node_list_org.append(node)
    # print(node_list_org)
    # print(len(node_list_org))
    # G = nx.Graph()
    # G.add_nodes_from(list(range(len(node_list_org))))
    # data = nx.to_pandas_edgelist(G_org)
    # ary_data = np.array(data)  # np.ndarray()
    # edge_list_org = ary_data.tolist()
    # edge_list = []
    # for edge in edge_list_org:
    #     edge_list.append((node_list_org.index(edge[0]), node_list_org.index(edge[1])))
    # node_num = len(G.nodes)
    # noise_edge = gen_noisy_input(ni_p, edge_list, node_num)
    #
    # edge_list_copy = copy.deepcopy(edge_list)
    # edge_list_copy.extend(noise_edge)
    # print(type(edge_list_copy))
    # G.add_edges_from(edge_list_copy)
    # theta = cal_dissimilarity(G, node_num)
    # # print('must')
    # # for e in edge_list:
    # #     print(theta[e[0]][e[1]])
    #
    # for dolphins:'gt'
    # football_ground_truth = [G_org.nodes[v]['value'] for v in G_org.nodes()]
    # ground_truth_index = get_ground_truth(football_ground_truth)
    ground_truth_dict = {i: [ground_truth_index[i]] for i in range(node_num)}
    ground_truth_com = node_com_to_com_node(ground_truth_dict)
    ground_truth_list = [set(v) for v in ground_truth_com.values()]
    print(ground_truth_com)
    theta = cal_dissimilarity(G, node_num)
    must_link, cannot_link = gen_prior_info(mu_p, cn_p, edge_list, ground_truth_com)
    noisy_must_link, noisy_cannot_link = gen_noisy_prior_info(mu_p, cn_p, np_p, edge_list, node_num, ground_truth_com,
                                                              theta)
    noisy_must_link, noisy_cannot_link = clean_noise(noisy_must_link, noisy_cannot_link, theta)
    # print('cannot')
    # for e in cannot_link:
    #     print(theta[e[0]][e[1]])
    # football_ground_truth_list = [7, 0, 2, 3, 7, 3, 2, 8, 8, 7, 3, 10, 6, 2, 6, 2, 7, 9, 6, 1, 9, 8, 8, 7, 10, 0, 6, 9,
    #                               11, 1, 1, 6, 2, 0, 6, 1, 5, 0, 6, 2, 3, 7, 5, 6, 4, 0, 11, 2, 4, 11, 10, 8, 3, 11, 6,
    #                               1, 9, 4, 11, 10, 2, 6, 9, 10, 2, 9, 4, 11, 8, 10, 9, 6, 3, 11, 3, 4, 9, 8, 8, 1, 5, 3,
    #                               5, 11, 3, 6, 4, 9, 11, 0, 5, 4, 4, 7, 1, 9, 9, 10, 3, 6, 2, 1, 3, 0, 7, 0, 2, 3, 8, 0,
    #                               4, 8, 4, 9, 11]
    # football_ground_truth_dict = {vertex: 0 for vertex in range(node_num)}
    # #-----------------end football---------------------

    # node_num, edge_list, ground_truth_lfr_org = get_lfr()
    # ground_truth_lfr = node_com_to_com_node(ground_truth_lfr_org)

    node_list = list(range(node_num))
    edge_num = len(edge_list)

    # ------------lfr--------------
    # G = nx.Graph()
    # G.add_nodes_from(node_list)
    # G.add_edges_from(edge_list)
    # ------------end lfr--------------
    '''
    for i in range(0, 2):
        for j in range(len(data)):
            if data[i][j] not in node_list:
                node_list.append(data[i][j])
                if data[i][j] > m_temp:
                    m_temp = data[i][j] 
    '''

    # ----------------initialization: get_edge()----------------
    edge_dict, edge_index_dict = get_edge(edge_num, node_num, edge_list)  # [[a1,b1],[a2,b2]]
    # ----get neighbor dict-------
    neighbor_dict = {node: [] for node in range(node_num)}
    degree_dict = {node: [] for node in range(node_num)}
    for node, value in edge_dict.items():
        neighbor_dict[node] = set(list(itertools.chain.from_iterable(value)))
        degree_dict[node] = len(value)
    print("dict")

    centrality_dict = nx.eigenvector_centrality_numpy(G)  # cal_eigenvector_centrality(node_list, edge_list)
    m = len(edge_list)
    print("centrality")
    neighbor_edge_dict = find_neighbor_edge(edge_list, edge_index_dict)
    print("neighbor")
    pop_size = 30
    popu = Population(pop_size)
    print("population")

    M = nx.convert_matrix.to_numpy_matrix(G)
    A = M.getA()

    init_list = init(G, m, popu, degree_dict, centrality_dict, node_list, edge_list, neighbor_dict, neighbor_edge_dict,
                     A, pop_size,
                     must_link, cannot_link)
    popu.set_individual_list(init_list)

    last_fitness = 0
    # Termination test
    counter = 0
    flag = 1
    # parameters
    max_iter = 2
    c_p = 0.5  # karate:0.5  # 0.15  # 0.1
    m_p = 0.05  # 0.1
    marginal_p = 0.4  # karate:0.5  # 0.01
    lsp = 0.1

    popu.update_fitness_list()
    print("beginning")
    popu.update_diversity()
    initial_diversity = popu.get_diversity()
    print("initial_diversity", initial_diversity)
    groundTruth = [0, 0, 0, 0, 0, 0, 0, 0, 1, 1, 0, 0, 0, 0, 1, 1, 0, 0, 1, 0, 1, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
    print("start")
    while (counter < max_iter):
        function1_values = copy.deepcopy(popu.get_fitness_list())
        function2_values = copy.deepcopy(popu.get_fitness_list2())
        non_dominated_sorted_solution = fast_non_dominated_sort(function1_values[:], function2_values[:])
        print("The best front for Generation number ", counter, " is")
        for valuez in non_dominated_sorted_solution[0]:  # index
            print(round(popu.get_fitness_list()[valuez], 3), end=" ")
        print("\n")
        crowding_distance_values = []
        for i in range(0, len(non_dominated_sorted_solution)):
            crowding_distance_values.append(
                crowding_distance(function1_values[:], function2_values[:], non_dominated_sorted_solution[i][:]))
        individual_list2 = copy.deepcopy(popu.get_individual_list())
        individual_list = copy.deepcopy(popu.get_individual_list())
        # individual_list_2
        # Generating offsprings
        while (len(individual_list2) != 2 * pop_size):
            a1 = random.randint(0, pop_size - 1)
            b1 = random.randint(0, pop_size - 1)
            parent1 = individual_list[a1]
            parent2 = individual_list[b1]
            child1, child2 = crossover(m, c_p, parent1, parent2)
            child1 = mutate(edge_list, child1, m_p, neighbor_edge_dict, marginal_p, degree_dict, node_num, m, A)
            child2 = mutate(edge_list, child2, m_p, neighbor_edge_dict, marginal_p, degree_dict, node_num, m, A)
            com_edge1, com_index1 = decode(child1, m)
            com_edge2, com_index2 = decode(child2, m)
            child1 = after_decoding_2(child1, com_index1, neighbor_edge_dict)
            child2 = after_decoding_2(child2, com_index2, neighbor_edge_dict)
            com_edge1, com_index1 = decode(child1, m)
            com_edge2, com_index2 = decode(child2, m)
            node_com_1 = com_edge_to_node(len(node_list), com_index1, edge_list)
            com_node_1 = node_com_to_com_node(node_com_1)
            edge_com1 = edge_com_to_com_edge(com_edge1)
            edge_com2 = edge_com_to_com_edge(com_edge2)
            c1_fitness = calc_EQ(m, degree_dict, com_node_1, A, node_com_1)
            c1_fitness2 = cal_fitness2(noisy_must_link, noisy_cannot_link, node_com_1, centrality_dict)

            node_com_2 = com_edge_to_node(len(node_list), com_index2, edge_list)
            com_node_2 = node_com_to_com_node(node_com_2)

            c2_fitness = calc_EQ(m, degree_dict, com_node_2, A, node_com_2)
            c2_fitness2 = cal_fitness2(noisy_must_link, noisy_cannot_link, node_com_2, centrality_dict)
            c1 = Individual(genes=child1, community_edge=com_edge1, community_by_index=com_index1,
                            com_node=com_node_1)
            c2 = Individual(genes=child2, community_edge=com_edge2, community_by_index=com_index2,
                            com_node=com_node_2)
            c1.set_fitness(c1_fitness)
            c1.set_fitness2(c1_fitness2)
            c2.set_fitness(c2_fitness)
            c2.set_fitness2(c2_fitness2)
            individual_list2.append(c1)
            individual_list2.append(c2)

        function1_values2 = [individual_list2[i].get_fitness() for i in range(0, 2 * pop_size)]
        function2_values2 = [individual_list2[i].get_fitness2() for i in range(0, 2 * pop_size)]
        non_dominated_sorted_solution2 = fast_non_dominated_sort(function1_values2[:], function2_values2[:])
        crowding_distance_values2 = []
        for i in range(0, len(non_dominated_sorted_solution2)):
            crowding_distance_values2.append(
                crowding_distance(function1_values2[:], function2_values2[:], non_dominated_sorted_solution2[i][:]))
        new_solution = []
        for i in range(0, len(non_dominated_sorted_solution2)):
            non_dominated_sorted_solution2_1 = [
                index_of(non_dominated_sorted_solution2[i][j], non_dominated_sorted_solution2[i]) for j in
                range(0, len(non_dominated_sorted_solution2[i]))]
            front22 = sort_by_values(non_dominated_sorted_solution2_1[:], crowding_distance_values2[i][:])
            front = [non_dominated_sorted_solution2[i][front22[j]] for j in
                     range(0, len(non_dominated_sorted_solution2[i]))]
            front.reverse()
            front_best = copy.deepcopy(front)
            for value in front:
                new_solution.append(value)
                if (len(new_solution) == pop_size):
                    break
            if (len(new_solution) == pop_size):
                break
        individual_list_temp = [individual_list2[i] for i in new_solution]
        popu.set_individual_list(individual_list_temp)
        popu.update_fitness_list()
        counter = counter + 1
    non_dominated_sorted_solution = fast_non_dominated_sort(popu.get_fitness_list(), popu.get_fitness_list2())

    # for sol in non_dominated_sorted_solution:
    #     for i in sol:
    #         draw(node_list, edge_list, popu.get_individual_list()[i])
    G_final = nx.Graph()
    G_final.add_nodes_from(node_list)
    G_final.add_edges_from(edge_list)
    function1 = [popu.get_fitness_list()[i] * -1 for i in non_dominated_sorted_solution[0]]
    function2 = [popu.get_fitness_list2()[i] * -1 for i in non_dominated_sorted_solution[0]]

    onmi_list = []
    omega = []
    count_print = 0
    for individual in popu.get_individual_list():
        com_node = individual.get_com_node()
        com_node_list = [set(v) for v in com_node.values()]
        onmi_list.append(onmi.onmi(com_node_list, ground_truth_list, variant="MGH"))
        print("Q", individual.get_fitness())
        print("onmi", onmi_list[count_print])
        omega.append(evaluate(ground_truth_com, com_node))
        count_print += 1
        draw_result(node_com_to_com_node(com_node), G)
    write_all(popu.get_individual_list(), onmi_list, omega)
    plt.xlabel('Function 1', fontsize=15)
    plt.ylabel('Function 2', fontsize=15)
    plt.scatter(function1, function2)
    plt.show()

if __name__ == '__main__':
    main()
